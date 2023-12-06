from datetime import datetime, timedelta
import pytz
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side

from selenium import webdriver

import re
import os
import glob
import openpyxl
import pandas as pd
import pdb

def cp_from_gs():
    GS_PATH = "gs://wtc_save/data/*"
    BASE_DATA = os.path.join(os.getcwd(), "data/")
    os.system(f"gcloud storage cp -r '{GS_PATH}' '{BASE_DATA}'")

def cp_to_gs():
    GS_PATH = "gs://wtc_save/data/"
    BASE_DATA = os.path.join(os.getcwd(), "data/*")
    os.system(f"gcloud storage cp -r '{BASE_DATA}' '{GS_PATH}'")
    
def extract_kakutei(string):
    # 予約番号を抽出
    print(f"{string=}")
    facility_pattern = r"施設名\n(.*?)\n"
    reservation_date_pattern = r"予約日\n(.*?)\n"
    usage_time_pattern = r"使用時間\n(.*?)\n"
    lottery_date_pattern = r"抽選日\n(.*?)\n"
    reservation_number_pattern = r"予約番号\n([0-9]*)$"
    
    facility_match = re.search(facility_pattern, string)
    reservation_date_match = re.search(reservation_date_pattern, string)
    usage_time_match = re.search(usage_time_pattern, string)
    lottery_date_match = re.search(lottery_date_pattern, string)
    reservation_number_match = re.search(reservation_number_pattern, string)
    
    if facility_match:
        facility_name = facility_match.group(1)
    else:
        facility_name = None
    
    if reservation_date_match:
        reservation_date = reservation_date_match.group(1)
    else:
        reservation_date = None
    
    if usage_time_match:
        usage_time = usage_time_match.group(1)
    else:
        usage_time = None
    
    if lottery_date_match:
        lottery_date = lottery_date_match.group(1)
    else:
        lottery_date = None
    
    if reservation_number_match:
        reservation_number = reservation_number_match.group(1)
    else:
        reservation_number = None

    print(f"施設名: {facility_name}")
    print(f"予約日: {reservation_date}")
    print(f"使用時間: {usage_time}")
    print(f"抽選日: {lottery_date}")
    print(f"予約番号: {reservation_number}")


    return {"court":facility_name,
            "date": reservation_date,
            "time_range": usage_time, 
            "reservation_number": int(reservation_number)
           }
    

def save_to_excel(df, file_name):
    # Create a workbook and worksheet
    wb = Workbook()
    ws = wb.active

    # Set the title and apply formatting
    timezone_utc_plus_9 = pytz.timezone('Asia/Tokyo')
    time_scraped = datetime.now(timezone_utc_plus_9)
    str_time_scraped = time_scraped.strftime('%m月%d日 %H:%M')
    title = f"埼玉県営テニスコート名義 {str_time_scraped}取得"
    
    ws['A1'] = title
    title_font = Font(size=14, bold=True)
    ws['A1'].font = title_font

    # Write the DataFrame to the worksheet
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # Set column width for better readability
    column_widths = [15, 20, 20, 15]
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width
      
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal='left')

    # Write the total number of rows at the bottom
    total_rows = len(df) + 4  # Add 4 for the two additional rows
    ws.cell(row=total_rows-1, column=1, value="利用可名義数:")
    ws.cell(row=total_rows-1, column=2, value=len(df))

    ws.cell(row=total_rows, column=1, value="総票数:")
    ws.cell(row=total_rows, column=2, value=len(df)*4)

    # Save the Excel file
    wb.save(file_name)


def check_schedule_within_30_minutes():
    # 日本の時刻
    timezone_utc_plus_9 = pytz.timezone('Asia/Tokyo')
    current_time = datetime.now(timezone_utc_plus_9)

    # 毎週金曜日のスケジュール
    friday_schedule = {'day_of_week': 4, 'start_time': '03:00', 'end_time': '03:30'}

    # 毎月第2水曜日のスケジュール
    second_wednesday = current_time.replace(day=15)
    while second_wednesday.weekday() != 2:  # 水曜日を探す
        second_wednesday += timedelta(days=1)
    second_wednesday_schedule = {'day': second_wednesday.day, 'start_time': '22:30', 'end_time': '23:59'}
    next_day_schedule = {'day': (second_wednesday + timedelta(days=1)).day, 'start_time': '00:00', 'end_time': '08:00'}

    # 現在時刻から30分後までをチェック
    for _ in range(31):
        # Check the current time and two schedules
        if check_single_schedule(current_time, friday_schedule) or check_single_schedule(current_time, second_wednesday_schedule) or check_single_schedule(current_time, next_day_schedule):
            return 1
        # Increment current_time by 1 minute for the next iteration
        current_time += timedelta(minutes=1)

    return 0


def check_single_schedule(current_time, schedule):
    start_time = datetime.strptime(schedule['start_time'], '%H:%M').time()
    end_time = datetime.strptime(schedule['end_time'], '%H:%M').time()

    if not (day := schedule.get("day")):
        # 毎週金曜日について
        day_of_week = schedule['day_of_week']

        if current_time.weekday() == day_of_week and start_time <= current_time.time() <= end_time:
            return True
    else:
        # 第2水曜日の判定
        if current_time.day == day and start_time <= current_time.time() <= end_time:
            return True

    return False

import pandas as pd

def create_date_dict(data):
    rows_dict = dict()
    columns = ['08:30', '10:30', '12:30', '14:30', '16:30', '18:30']
    
    for court, value in data.items():
        for date, times in value.items():
            rows_dict[date] = []
        
    # Iterate through the data dictionary to populate the DataFrame rows
    for court, value in data.items():
        for date, times in value.items():
            rows_dict[date].append([int(court)+1] + [times[time] for time in columns])
    return rows_dict

def save_votes(data, file_path):
    date_dict = create_date_dict(data)
    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

    # Create a workbook and worksheet
    wb = Workbook()
    ws = wb.active

    # Set the title and apply formatting
    title = "票数"
    ws['A1'] = title
    title_font = Font(size=14, bold=True)
    ws['A1'].font = title_font
    
    count = 0
    for i, (date, courts) in enumerate(date_dict.items()):
        # set border to each cell
        for row in ws[f'B{1+i*16}:I{1+i*16+12}']:
            for cell in row:
                cell.border = thin_border
        ws.cell(row=1+i*16, column=2, value= f"{date}")
        # create header
        start_times = ['08:30', '10:30', '12:30', '14:30', '16:30', '18:30'] 
        for j, value in enumerate(start_times):
            ws.cell(row=1+i*16, column=2+j+1, value= f"{value}")

        # write number of votes for each day
        for j, court in enumerate(courts):
            ws.cell(row=1+i*16+(j+1), column=2, value= f"{j+1}番")
            ws.cell(row=1+i*16+(j+1), column=2+len(court), value= f"{j+1}番")
            for k, vote in enumerate(court[1:]):
                if vote=="unavailable":
                    vote = "休"
                ws.cell(row=1+i*16+(j+1), column=2+(k+1), value= f"{vote}")

    # Save the Excel file
    wb.save(file_path)


def get_vote_dest():
    # gs://wtc_saveにアップロードしたものをコピー
    # cp_from_gs()
    # ファイルを検索するディレクトリのパスを指定
    DATA_BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
    
    # ファイルを検索
    files = glob.glob(os.path.join(DATA_BASE, '*最終形態*'))
    
    # ファイルが見つからなかった場合の処理
    if not files:
        print("最終形態と書かれたファイルが見つかりませんでした")
    else:
        # ファイルの更新日時でソートし、最新のファイルを取得
        latest_file = max(files, key=os.path.getmtime)
        print("最新の最終形態ファイルのパス:", latest_file)
    df_path = latest_file

    wb = openpyxl.load_workbook(df_path)
    ws = wb.worksheets[0]

    valid_row_count = 0
    for row in ws.iter_rows():  # This will loop through all rows in the worksheet
      valid_row_count += 1
    # print(f"{valid_row_count=}")

    vote_dest = pd.DataFrame(columns= ["date","time","court"])
    row_idx = 1
    n =0
    while 1+(row_idx-1)*16 < valid_row_count:
      date = ws[f"B{1+(row_idx-1)*16}"].value
      # print(date)
      if date == None:
        row_idx += 1
        continue
      # 時間帯を取得
      times = [time.value for time in ws[f'C{1+(row_idx-1)*16}:I{1+(row_idx-1)*16}'][0]]
      # 1-12番コートの票数
      for i, row in enumerate(ws[f'B{2+(row_idx-1)*16}:I{1+(row_idx-1)*16+12}']):
        court = i+1
        for time, cell in zip(times, row):
          if cell.value is None:
              cell.value = 0
          if cell.value == "休" or (rep := int(cell.value)) == 0:
            continue
          to_add = pd.DataFrame({"date": [date]*rep,"time": [time]*rep,"court": [court]*rep})
          vote_dest = pd.concat([vote_dest, to_add])
          n += rep
      row_idx+=1
    vote_dest = vote_dest.reset_index()
    vote_dest = vote_dest.drop("index", axis=1)

    return vote_dest

def flatten_accounts(accounts):
  use_accounts = []
  for _, account in accounts.iterrows():
    use_accounts.extend([account["通し番号"] for _ in range(int(account["n_votes"]))])
  return use_accounts

def append_accounts_to_vote_dest(vote_dest):
    """vote_destに、どのアカウントを用いて投票するかの情報を付け加える"""
    
    DATA_BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
    
    # ファイルを検索
    files = glob.glob(os.path.join(DATA_BASE, '埼玉県営利用可名義*.xlsx*'))

    # ファイルが見つからなかった場合の処理
    if not files:
        print("最終形態と書かれたファイルが見つかりませんでした")
    else:
        # ファイルの更新日時でソートし、最新のファイルを取得
        latest_file = max(files, key=os.path.getmtime)
        print("最新の利用可能名義ファイルのパス:", latest_file)

    # dfの処理
    df = pd.read_excel(latest_file, header=1)
    df = df[df["通し番号"] != "利用可名義数:"]
    df = df[df["通し番号"] != "総票数:"]

    # accountsの処理
    accounts = df
    accounts["n_votes"] = 4
    
    d = flatten_accounts(accounts)
    c = []
    while len(d) > 0:
      b = set(d)
      for i in b:
        d.remove(i)
        c.append(i)

    vote_dest["account"] = pd.Series(c)
    return vote_dest

def save_won_votes(votes_won):
    votes_won = votes_won.drop_duplicates()
    data =  votes_won

    # コート名を変換する辞書
    court_dict = {
        '第１テニスコート第１クレーコート': 1,
        '第１テニスコート第２クレーコート': 2,
        '第１テニスコート第５人工芝コート': 5,
        '第１テニスコート第８人工芝コート': 8,
        '第１テニスコート第３人工芝コート': 3,
        '第１テニスコート第４人工芝コート': 4,
        '第１テニスコート第７人工芝コート': 7,
        '第１テニスコート第６人工芝コート': 6,
        '第１テニスコート第９人工芝コート': 9,
        '第２テニスコート第１１人工芝コート': 11,
        '第１テニスコート第１０人工芝コート': 10,
        '第２テニスコート第１２人工芝コート': 12,
    }
    
    # 日付の形式変換（'2023/09/12' → '09-12'）
    data['date'] = pd.to_datetime(data['date']).dt.strftime('%m-%d')
    
    # time_rangeの形式変換（'10:30～12:30' → '10:30'）
    data['time_range'] = data['time_range'].str.split('～').str[0]
    
    # コート名の変換
    data['court'] = data['court'].map(court_dict)
    votes_won = data

    date_list = []
    for r in votes_won["date"].unique():
      date_list.append(r)
    date_list.sort()

    thin_border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    
    # Create a workbook and worksheet
    wb = Workbook()
    ws = wb.active
    
    # Set the title and apply formatting
    title = "埼玉県営コート　当選票の通し番号"
    ws['A1'] = title
    title_font = Font(size=14, bold=True)
    ws['A1'].font = title_font
    
    for i, date in enumerate(date_list):
      # set border to each cell
      for row in ws[f'A{1+i*16+1}:H{1+i*16+12+1}']:
        for cell in row:
          cell.border = thin_border
          cell.alignment = Alignment(horizontal="left")
      ws.cell(row=1+i*16+1, column=1, value= f"{date}")
      # create header
      start_times = ['08:30', '10:30', '12:30', '14:30', '16:30', '18:30']
      for j, value in enumerate(start_times):
        ws.cell(row=1+i*16+1, column=1+j+1, value= f"{value}")
    
      # write number of votes for each day\
      for j in range(1, 13):
        ws.cell(row=1+i*16+j+1, column=1, value= f"{j}番")
        ws.cell(row=1+i*16+j+1, column=1+len(start_times)+1, value= f"{j}番")
        for k, time in enumerate(start_times):
          try:
            num = list(votes_won[(votes_won["time_range"] == time) & (votes_won["date"] == date) & (votes_won["court"] == j)]["通し番号"])[0]
            ws.cell(row=1+i*16+j+1, column=1+(k+1), value= f"{num}")
          except IndexError:
            pass

    DATA_BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
    wb.save(os.path.join(DATA_BASE, "votes_won.xlsx"))
    # cp_to_gs()
    
def get_driver():
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')

    driver = webdriver.Remote(
     command_executor='http://localhost:4444/wd/hub',
     options = options,
   )
    return driver
def excel_serial_date_to_date(serial_date):
    # Excelの基準日
    import datetime
    excel_start_date = datetime.date(1899, 12, 30)
    
    # Excelのバグ（1900年をうるう年と誤認）のための補正
    if serial_date > 59:
        serial_date -= 1

    # 実際の日付を計算
    return excel_start_date + datetime.timedelta(days=int(serial_date))
